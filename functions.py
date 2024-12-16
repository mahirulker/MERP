import pyodbc
import pandas as pd
import openpyxl
import os
import json
import win32com.client as win32
import pythoncom
import streamlit as st
from datetime import datetime
import win32gui
import win32con
import win32api
import time

# Veritabanı bağlantı fonksiyonu
def get_db_connection():
    conn = pyodbc.connect(
        'DRIVER={SQL Server};'
        'SERVER=DESKTOP-VT3UCMO\\SQLEXPRESS;'
        'DATABASE=TEST_DB;'
        'Trusted_Connection=yes;'
    )
    return conn

# Excel ve PDF işlemleri
def excel_to_pdf(excel_path):
    try:
        pythoncom.CoInitialize()
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # PDF klasör yolu
        pdf_klasoru = r"C:\Users\ACPE\Desktop\cursor\teklif_web_test\teklif_pdf"
        
        # PDF klasörünü kontrol et ve oluştur
        if not os.path.exists(pdf_klasoru):
            os.makedirs(pdf_klasoru)
        
        # Excel dosyasının tam yolu
        abs_excel_path = os.path.abspath(excel_path)
        
        # PDF dosya adını oluştur
        pdf_adi = os.path.basename(excel_path).replace('.xlsx', '.pdf')
        pdf_path = os.path.join(pdf_klasoru, pdf_adi)
        
        # Excel'i aç ve PDF'e çevir
        wb = excel.Workbooks.Open(abs_excel_path)
        
        try:
            wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
        except:
            sheet = wb.ActiveSheet
            sheet.ExportAsFixedFormat(0, pdf_path)
        
        wb.Close(False)
        excel.Quit()
        
        return pdf_path
        
    except Exception as e:
        st.error(f"PDF dönüşüm hatası: {str(e)}")
        return None
    finally:
        try:
            if 'excel' in locals():
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

def create_excel(teklif_ozeti, musteri, tarih, cell_mapping):
    try:
        df = pd.DataFrame(teklif_ozeti)
        
        # Teklifler klasörünü kontrol et ve oluştur
        teklifler_klasoru = r"C:\Users\ACPE\Desktop\cursor\teklif_web_test\teklifler"
        if not os.path.exists(teklifler_klasoru):
            os.makedirs(teklifler_klasoru)
        
        # Dosya adı ve yolu oluştur
        dosya_adi = f"teklif_{musteri}_{tarih}.xlsx"
        kayit_yolu = os.path.join(teklifler_klasoru, dosya_adi)
        
        # Şablon dosyasını yükle
        wb = openpyxl.load_workbook(f"teklif_formlari/{musteri}_form.xlsx")
        ws = wb.active
        
        start_row = int(cell_mapping['urun_adi'][1:]) + 1
        
        for index, row in df.iterrows():
            current_row = start_row + index
            ws[f"{cell_mapping['urun_adi'][0]}{current_row}"].value = row['Ürün']
            ws[f"{cell_mapping['miktar'][0]}{current_row}"].value = row['Miktar']
            ws[f"{cell_mapping['birim_fiyat'][0]}{current_row}"].value = row['Birim Fiyat']
            ws[f"{cell_mapping['toplam'][0]}{current_row}"].value = row['Toplam']

        # Kaydet
        wb.save(kayit_yolu)
        
        return kayit_yolu if os.path.exists(kayit_yolu) else None

    except Exception as e:
        st.error(f"Excel oluşturma hatası: {str(e)}")
        return None

# Dosya işlemleri
def check_teklif_form(musteri):
    file_path = f"teklif_formlari/{musteri}_form.xlsx"
    return os.path.exists(file_path)

def save_teklif_form(musteri, file):
    os.makedirs("teklif_formlari", exist_ok=True)
    file_path = f"teklif_formlari/{musteri}_form.xlsx"
    with open(file_path, "wb") as f:
        f.write(file.getbuffer())

def delete_teklif_form(musteri):
    file_path = f"teklif_formlari/{musteri}_form.xlsx"
    if os.path.exists(file_path):
        os.remove(file_path)

def save_cell_mapping(musteri, cell_mapping):
    os.makedirs("cell_mappings", exist_ok=True)
    with open(f"cell_mappings/{musteri}_mapping.json", "w") as f:
        json.dump(cell_mapping, f)

def load_cell_mapping(musteri):
    try:
        with open(f"cell_mappings/{musteri}_mapping.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return None

# E-posta işlemleri
def send_email_with_outlook(to_address, subject, body, attachment_path=None):
    try:
        pythoncom.CoInitialize()
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to_address
        mail.Subject = subject
        mail.Body = body
        
        if attachment_path:
            abs_path = os.path.abspath(attachment_path)
            if os.path.exists(abs_path):
                mail.Attachments.Add(abs_path)
            else:
                st.error(f"Dosya bulunamadı: {abs_path}")
                return False
        
        mail.Send()
        return True
        
    except Exception as e:
        st.error(f"E-posta gönderim hatası: {str(e)}")
        return False
    finally:
        pythoncom.CoUninitialize()

def print_teklif(excel_path):
    try:
        pythoncom.CoInitialize()
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = True
        
        # Excel dosyasının tam yolu
        abs_excel_path = os.path.abspath(excel_path)
        
        # Excel'i aç
        wb = excel.Workbooks.Open(abs_excel_path)
        
        # Kısa bir bekleme süresi
        time.sleep(1)
        
        # Alt + F tuşlarını gönder
        win32api.keybd_event(0x12, 0, 0, 0)  # ALT tuşu basılı
        win32api.keybd_event(0x50, 0, 0, 0)  # P tuşu basılı
        win32api.keybd_event(0x50, 0, win32con.KEYEVENTF_KEYUP, 0)  # P tuşu bırakıldı
        win32api.keybd_event(0x12, 0, win32con.KEYEVENTF_KEYUP, 0)  # ALT tuşu bırakıldı
        
        return True
        
    except Exception as e:
        st.error(f"Yazdırma hatası: {str(e)}")
        return False
    finally:
        pythoncom.CoUninitialize()